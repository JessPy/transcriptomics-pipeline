from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

DEFAULT_ACCESSION_HEADERS = {
    "sample accession",
    "accession",
    "run accession",
    "sra accession",
    "experiment accession",
}


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _read_text_table(source: Path) -> tuple[list[str] | None, list[dict[str, str]]]:
    text = source.read_text(encoding="utf-8-sig").strip()
    if not text:
        return None, []

    sample = text.splitlines()[0]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        has_header = csv.Sniffer().has_header(text)
        reader = csv.reader(text.splitlines(), dialect)
        rows = [row for row in reader if any(field.strip() for field in row)]
    except csv.Error:
        rows = [[line.strip()] for line in text.splitlines() if line.strip()]
        has_header = False

    if not rows:
        return None, []

    if has_header and len(rows) > 1:
        headers = [row.strip() for row in rows[0]]
        data_rows = rows[1:]
    elif len(rows[0]) == 1:
        headers = ["accession"]
        data_rows = rows
    else:
        headers = [row.strip() for row in rows[0]]
        data_rows = rows[1:]

    normalized_headers = [_normalize_header(header) for header in headers]
    return normalized_headers, [dict(zip(normalized_headers, row, strict=True)) for row in data_rows]


def _read_excel_table(source: Path) -> tuple[list[str] | None, list[dict[str, str]]]:
    if load_workbook is None:
        raise ImportError(
            "openpyxl não está instalado. Instale com 'poetry install' ou adicione openpyxl como dependência."
        )

    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [tuple(cell.value if cell.value is not None else "" for cell in row) for row in worksheet.iter_rows()]
    rows = [tuple(str(value).strip() for value in row) for row in rows if any(str(value).strip() for value in row)]

    if not rows:
        return None, []

    headers = [str(value) for value in rows[0]]
    normalized_headers = [_normalize_header(header) for header in headers]
    return normalized_headers, [dict(zip(normalized_headers, row, strict=True)) for row in rows[1:]]


def _find_accession_header(headers: Iterable[str], requested: str | None = None) -> str | None:
    if requested:
        requested_normalized = _normalize_header(requested)
        for header in headers:
            if header == requested_normalized:
                return header
        return None

    for header in headers:
        if header in DEFAULT_ACCESSION_HEADERS:
            return header
    return None


def read_sample_accessions(source: Path, accession_column: str | None = None) -> list[str]:
    source = Path(source)
    if not source.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {source}")

    suffix = source.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        headers, rows = _read_excel_table(source)
    else:
        headers, rows = _read_text_table(source)

    if headers is None or not rows:
        raise ValueError(f"Nenhum dado válido encontrado em: {source}")

    selected_header = _find_accession_header(headers, accession_column)
    if selected_header is None:
        raise ValueError(
            "Coluna de acesso não encontrada. Use --accession-column com o nome correto da coluna "
            f"(por exemplo, 'Sample Accession'). Cabeçalhos encontrados: {headers}."
        )

    accessions = [row[selected_header].strip() for row in rows if row.get(selected_header)]
    accessions = [value for value in accessions if value]
    if not accessions:
        raise ValueError(
            f"Nenhuma accession encontrada na coluna '{selected_header}' do arquivo {source}."
        )
    return accessions
